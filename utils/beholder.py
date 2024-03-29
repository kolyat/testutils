#!/usr/bin/python3
# -*- coding: utf-8 -*-

import os
import time

from ppadb.client import Client
import openpyxl


DEVICE_IP = '192.168.0.100'
DEVICE_PORT = 5555
PACKAGE = 'com.android.systemui'

WORKBOOK = 'meminfo.xlsx'  # Full path

TIME = 20  # Minutes of total working time
DELAY = 3  # Seconds between requests


def ts():
    return int(time.time())


if __name__ == '__main__':
    # Start server and connect to device
    os.system('adb start-server')
    client = Client(host='127.0.0.1', port=5037)
    print(f'Connecting to {DEVICE_IP}:{DEVICE_PORT}... ', end='')
    if client.remote_connect(DEVICE_IP, DEVICE_PORT):
        print('OK')
    else:
        print('Failed')
        print(f'Unable to connect to {DEVICE_IP}')
        exit(1)
    device = client.device(f'{DEVICE_IP}:{DEVICE_PORT}')

    # Prepare workbook
    if os.path.isfile(WORKBOOK):
        wb = openpyxl.load_workbook(WORKBOOK)
        print(f'Open {WORKBOOK}')
    else:
        wb = openpyxl.Workbook()
        wb.save(WORKBOOK)
        print(f'Create {WORKBOOK}')
    ws = wb.create_sheet(str(ts()), 0)
    b1 = 'RAM used, K'
    row = 2

    # Collect data
    print(f'Starting to collect data for {PACKAGE}')
    print('{:>15}'.format(b1))
    wtime = TIME * 60
    while wtime > 0:
        raw = device.shell(f'dumpsys meminfo | grep {PACKAGE}')
        result = int(raw.strip().split('K')[0].replace(',', ''))
        ws.cell(row=row, column=1, value=result)
        wb.save(WORKBOOK)
        print('{:>15}'.format(result))
        time.sleep(DELAY)
        row += 1
        wtime -= DELAY
    print('All done.')

    # Close connection and exit
    wb.close()
    client.remote_disconnect()
    client.kill()
